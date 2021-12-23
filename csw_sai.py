import os
import copy
import openpyxl
from pathlib import Path
from itertools import combinations

from simpleai.search import (
    CspProblem,
    backtrack,
    HIGHEST_DEGREE_VARIABLE,
    MOST_CONSTRAINED_VARIABLE,
    LEAST_CONSTRAINING_VALUE,
)

HEAD_COUNSELORS = {'Brodi','Ceci','Ignacio','Jocelyn','John','Lily','Max','Micky','Noor'}
COUNSELORS = {'Alicja','Brent','Daniel','Dom','Gabriel','Ildar','Juliana','Murray','Nelson','Simone','Sparta','Sydy'}
INTERNS = {'Audrey','Ben','Jack','Jordan','Nathan','Sophie','Van'}
ARCHERY_CERT = {'Ignacio','Jocelyn','Dom','Gabriel','Simone','Sophie','Jack','Brent','Daniel'}
LIFEGUARD_CERT = {'Brodi','Ceci','Dom','Ildar','Sparta','Nathan','Van','Ben'}
# CELLS_TO_IGNORE = {14,17,60,77,80,153}
CELLS_TO_IGNORE = {12,15,58,75,78,151} # Original -2

def spreadsheet_tasks():
    # Get preferences spreadsheet
    preference_sheet_file = Path(os.path.dirname(os.path.abspath(__file__)),'data','Preferences 2022 CSW Staff.xlsx')
    preference_workbook_obj = openpyxl.load_workbook(preference_sheet_file)
    preference_sheet = preference_workbook_obj.active

    #Get tomorrow's schedule
    schedule_sheet_file = Path(os.path.dirname(os.path.abspath(__file__)),'data','Schedule.xlsx')
    schedule_workbook_obj = openpyxl.load_workbook(schedule_sheet_file)
    schedule_sheet = schedule_workbook_obj.active    

    # Get list of staff
    staff = []
    for row in preference_sheet.iter_rows(min_row=3, max_row=30, max_col=1):
        for cell in row:
            staff.append(cell.value)

    # Get the full list of camp activities
    activities = []
    for col in preference_sheet.iter_cols(min_row=2, max_row=2, max_col=preference_sheet.max_column):
        for cell in col:
            if cell.value != None:
                activities.append(cell.value)
    activities.append('Free Period')

    # Get the list of activities from tomorrow's schedule

    def get_activity_periods(period, first_row, last_row):
        for row_index, row in enumerate(schedule_sheet.iter_rows(min_row=first_row, max_row=last_row, max_col=schedule_sheet.max_column)):
            for cell_index, cell in enumerate(row):
                if cell.value != None:
                    if row_index % 2 == 0:      # If row is even, it lists the activities
                        period.append([cell.value])
                    else:
                        period[cell_index].append(cell.value)
        # print(period)
        return period

    period_1 = []
    period_2 = []
    period_3 = []
    period_4 = []
    period_5 = []

    get_activity_periods(period_1,1,2)
    get_activity_periods(period_2,3,4)
    get_activity_periods(period_3,5,6)
    get_activity_periods(period_4,7,8)
    get_activity_periods(period_5,9,10)

    schedule_activities = [period_1, period_2, period_3, period_4, period_5]
    print(schedule_activities)

    # schedule_activities.append('Free Period')

    # Get preferences and creates a dictionary
    preferences = {}
    for row_index, row in enumerate(preference_sheet.iter_rows(min_row=3, max_row=30, min_col=2, max_col=169)):
        temp_preference_list = []
        for cell_index, cell in enumerate(row):
            if cell_index not in CELLS_TO_IGNORE:
                if cell.value != None:
                    # print(cell.value)
                    if cell.value.casefold() == "lead":
                        temp_preference_list.append(2)
                    elif cell.value.casefold() == "assist":
                        temp_preference_list.append(1)
                    else:
                        temp_preference_list.append(0)
                else: # If there is no preference set, a default 1 (ASSIST) will be set
                    temp_preference_list.append(1)
        preferences[staff[row_index]] = temp_preference_list
        # print(preferences)

    return staff, activities, preferences, schedule_activities

staff, activities, preferences, schedule_activities = spreadsheet_tasks()

# VARIABLES AND DOMAINS

problem_variables = staff
domains = {}
schedule_activities_without_archery = copy.deepcopy(schedule_activities)
schedule_activities_without_waterfront = copy.deepcopy(schedule_activities)
schedule_activities_without_archery_and_waterfront = copy.deepcopy(schedule_activities)
if 'Archery' in schedule_activities:
    schedule_activities_without_archery.remove('Archery')
    schedule_activities_without_archery_and_waterfront.remove('Archery')
if 'Kayaking' in schedule_activities:
    schedule_activities_without_waterfront.remove('Kayaking')
    schedule_activities_without_archery_and_waterfront.remove('Kayaking')
if 'Swimming' in schedule_activities:
    schedule_activities_without_waterfront.remove('Swimming')
    schedule_activities_without_archery_and_waterfront.remove('Swimming')

# print(schedule_activities)
# print(schedule_activities_without_archery)
# print(schedule_activities_without_waterfront)
# print(schedule_activities_without_archery_and_waterfront)

for variable in problem_variables:
    if variable in ARCHERY_CERT and variable in LIFEGUARD_CERT: # If staff member is BOTH an archery instructor AND a lifeguard
        domains[variable] = schedule_activities
        # print(variable,schedule_activities)
    else:
        if variable in ARCHERY_CERT:
            domains[variable] = schedule_activities_without_waterfront
            # print(variable,schedule_activities_without_waterfront)
        elif variable in LIFEGUARD_CERT:
            domains[variable] = schedule_activities_without_archery
            # print(variable,schedule_activities_without_archery)
        else:
            domains[variable] = schedule_activities_without_archery_and_waterfront
            # print(variable,schedule_activities_without_archery_and_waterfront)

# CONSTRAINTS

constraints = []

def staff_camper_ratio(variables, values):
    # 3:1 to 5:1 ratio
    pass

def not_just_interns(variables, values):
    # An activity can't be staffed with only interns. It has to have at least one counselor assigned to it.
    pass

def take_preferences_into_account(variables, values):
    pass

#print(preference_sheet["D2"].value)

problem = CspProblem(problem_variables, domains, constraints)
solution = backtrack(problem, variable_heuristic=HIGHEST_DEGREE_VARIABLE, value_heuristic=LEAST_CONSTRAINING_VALUE)

#print("Solution:")
#print(solution)